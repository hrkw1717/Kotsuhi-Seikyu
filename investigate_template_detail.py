import openpyxl

def detail_investigate(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    
    print(f"Sheet: {ws.title}")
    
    # 列幅の確認 (1/10インチ単位的なもの)
    print("\n--- Column Widths ---")
    for i in range(1, 11):
        col_letter = openpyxl.utils.get_column_letter(i)
        width = ws.column_dimensions[col_letter].width
        print(f"Col {col_letter}: {width}")
        
    # 行高さの確認
    print("\n--- Row Heights ---")
    for i in range(1, 15):
        height = ws.row_dimensions[i].height
        print(f"Row {i}: {height}")

    # フォントと配置の確認 (一部の代表的なセル)
    print("\n--- Font and alignment ---")
    target_cells = ["A1", "A2", "A8", "B8", "C8", "D8", "E8"]
    for coord in target_cells:
        cell = ws[coord]
        print(f"Cell {coord}: Font={cell.font.name}, Size={cell.font.sz}, Bold={cell.font.bold}, Align={cell.alignment.horizontal}")

detail_investigate("テンプレート.xlsx")
