import os
import datetime
import tkinter as tk
from tkinter import ttk, messagebox
import ctypes
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# 設定
TEMPLATE_PATH = r"C:\Users\sbs\Documents\Antigravity\Tokeidai\テンプレート.xlsx"
SHIFT_TABLE_PATH = r"C:\Users\sbs\Documents\Antigravity\Tokeidai\シフト表時計台警備通年.xlsx"
SAVE_DIR = r"C:\Users\sbs\Documents\Antigravity\Tokeidai"
NAME = "堀川"
BG_COLOR = "#D1F0BB" # 黄緑色 (Light Yellow Green)

def get_month_options():
    """先月、今月、来月の年月の選択肢を生成する"""
    today = datetime.date.today()
    options = []
    
    # 先月
    first_of_this_month = today.replace(day=1)
    last_month = first_of_this_month - datetime.timedelta(days=1)
    options.append(last_month.strftime("%Y年%m月").replace("年0", "年")) # 0埋めなし対応
    
    # 今月
    options.append(today.strftime("%Y年%m月").replace("年0", "年"))
    
    # 来月
    next_month = (first_of_this_month + datetime.timedelta(days=32)).replace(day=1)
    options.append(next_month.strftime("%Y年%m月").replace("年0", "年"))
    
    return options

def format_month_label(month_str):
    """'2026年1月' -> '1月' に変換"""
    return month_str.split("年")[1].lstrip("0")

def process_excel(selected_year_month):
    try:
        # Step 1: 年と月の抽出
        month_label = format_month_label(selected_year_month) # "1月" など
        month_num = int(month_label.replace("月", ""))
        
        # 保存ファイル名の設定 (Step 3)
        save_filename = f"交通費請求{month_num}月-{NAME}.xlsx"
        save_path = os.path.join(SAVE_DIR, save_filename)
        
        # Step 2: テンプレートを開く
        wb_template = openpyxl.load_workbook(TEMPLATE_PATH)
        ws_claim = wb_template["交通費請求用紙 "]
        
        # Step 4: A2セルに代入し、書式を設定
        ws_claim["A2"] = f"{selected_year_month}分"
        ws_claim["A2"].font = Font(size=18)
        ws_claim["A2"].alignment = Alignment(horizontal="center", vertical="center")
        
        # Step 7: シフト表を開く
        wb_shift = openpyxl.load_workbook(SHIFT_TABLE_PATH, data_only=True)
        
        # Step 8: シフト表の特定
        target_ws = None
        start_row = None
        
        # 全シートを閲覧
        for sheet_name in wb_shift.sheetnames:
            ws = wb_shift[sheet_name]
            # シート内で月ラベルを探す (C列付近にあることが多い)
            for row in range(1, 100): # 上方100行程度を捜索
                for col in range(1, 10): # A-I列程度を捜索
                    val = ws.cell(row=row, column=col).value
                    if val == month_label:
                        target_ws = ws
                        start_row = row
                        break
                if target_ws:
                    break
            if target_ws:
                break
        
        if not target_ws:
            raise Exception(f"シフト表の中に「{month_label}」が見つかりませんでした。")

        # Step 10: 堀川さんの行を探す
        horikawa_row = None
        # 月ラベルの数行下から名前を探す
        for r in range(start_row, start_row + 20):
            for c in range(1, 5):
                if ws.cell(row=r, column=c).value == NAME:
                    horikawa_row = r
                    break
            if horikawa_row:
                break
        
        if not horikawa_row:
            raise Exception(f"シート「{target_ws.title}」の中に「{NAME}」が見つかりませんでした。")

        # シフト表の日付列（月ラベルの行の右側）を特定
        # 調査結果によると、月ラベルの右側(D列以降)に1, 2, 3...と並んでいる
        day_cols = {}
        for c in range(1, target_ws.max_column + 1):
            val = target_ws.cell(row=start_row, column=c).value
            if isinstance(val, (int, float)) or (isinstance(val, str) and val.isdigit()):
                day_cols[int(val)] = c

        # Step 14: 交通費請求用紙のループ (A8セルが1日の行、7行目がヘッダー)
        # テンプレートは A8:1, A9:2 ... と日付が入っているはず
        for claim_row in range(8, 40): # 8行目から1ヶ月分
            day_val = ws_claim.cell(row=claim_row, column=1).value
            if not isinstance(day_val, (int, float)):
                continue
            
            day = int(day_val)
            if day not in day_cols:
                continue
                
            # Step 11: シフトが「出」か確認
            shift_status = target_ws.cell(row=horikawa_row, column=day_cols[day]).value
            
            # Step 12: 「出」なら記入
            if shift_status == "出":
                ws_claim.cell(row=claim_row, column=2).value = "MMS"      # 得意先名
                ws_claim.cell(row=claim_row, column=3).value = "時計台"    # 現場名
                ws_claim.cell(row=claim_row, column=4).value = "東屯田通～西4丁目～東屯田通" # 往復移動経路
                ws_claim.cell(row=claim_row, column=5).value = 460        # 運賃 (E列)
                # テンプレートの構造を確認したところ E列(5列目)が運賃
        
        # Step 3: 別名で保存
        wb_template.save(save_path)
        return save_path

    except Exception as e:
        raise e

def run_app():
    # コンソールウィンドウを最小化する (Windowsのみ)
    try:
        hwnd = ctypes.windll.kernel32.GetConsoleWindow()
        if hwnd:
            ctypes.windll.user32.ShowWindow(hwnd, 6) # 6 = SW_MINIMIZE
    except Exception:
        pass

    root = tk.Tk()
    root.title("交通費請求書作成")
    root.geometry("400x300")
    root.configure(bg=BG_COLOR)

    # スタイルの設定 (Comboboxなどの背景色対応)
    style = ttk.Style()
    style.configure("TCombobox", fieldbackground="white")

    label = tk.Label(root, text="年月を指定してください。", bg=BG_COLOR, font=("MS Gothic", 11, "bold"))
    label.pack(pady=20)

    options = get_month_options()
    selected_var = tk.StringVar(value=options[0]) # 先月をデフォルト
    
    combo = ttk.Combobox(root, textvariable=selected_var, values=options, state="readonly", width=20)
    combo.pack(pady=10)

    def on_click():
        month = selected_var.get()
        try:
            path = process_excel(month)
            messagebox.showinfo("成功", f"作成しました：\n{path}")
            
            # 作成したファイル、シフト表、およびフォルダを開く (Windows)
            os.startfile(path)              # 今保存したファイル
            os.startfile(SHIFT_TABLE_PATH)  # シフト表
            os.startfile(SAVE_DIR)          # 保存先フォルダ
            
            root.destroy()
        except Exception as e:
            messagebox.showerror("エラー", str(e))

    # ボタンサイズを大きくする (paddingを増やす)
    btn = tk.Button(root, text="実 行", command=on_click, font=("MS Gothic", 12, "bold"), 
                    padx=40, pady=20, bg="#FFFFFF")
    btn.pack(pady=30)

    root.mainloop()

if __name__ == "__main__":
    run_app()
