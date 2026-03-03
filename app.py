import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from streamlit_gsheets import GSheetsConnection
import datetime
from datetime import timezone, timedelta
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.utils import formataddr
import tempfile
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from io import BytesIO
from pypdf import PdfReader, PdfWriter
import calendar
import base64
import fitz # PyMuPDF
import time
from dotenv import load_dotenv

# .env ファイルがあれば読み込む（ローカル用）
load_dotenv()

# --- 定数設定 ---
APP_TITLE = "時計台警備 交通費請求ツール"
VALID_IDS = ["yama", "saka", "hori"]
SHARED_PASSWORD = "tokei"

# 本番環境（Streamlit Cloud）では st.secrets から取得します
# ローカル環境では環境変数（またはデフォルト値）を使用します
try:
    COMPANY_EMAIL_DEFAULT = st.secrets.get("COMPANY_EMAIL_DEFAULT", "sbs@sobun.net")
    SENDER_EMAIL = st.secrets.get("SENDER_EMAIL", os.getenv("SENDER_EMAIL", "horikawa1717@gmail.com"))
    EMAIL_PASSWORD = st.secrets.get("EMAIL_PASSWORD", os.getenv("EMAIL_PASSWORD", ""))
except:
    COMPANY_EMAIL_DEFAULT = os.getenv("COMPANY_EMAIL_DEFAULT", "sbs@sobun.net")
    SENDER_EMAIL = os.getenv("SENDER_EMAIL", "horikawa1717@gmail.com")
    EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD", "")

def get_target_company_email():
    """ログイン中のユーザーの設定から会社メアドを取得し、なければデフォルトを返す"""
    if "user_id" not in st.session_state:
        return COMPANY_EMAIL_DEFAULT
    
    df_all = load_mypage()
    user_row = df_all[df_all["ID"] == st.session_state.user_id]
    if not user_row.empty:
        val = user_row.iloc[0].get("会社メアド", "")
        if val and not (isinstance(val, float) and pd.isna(val)):
            return val
    return COMPANY_EMAIL_DEFAULT

def render_global_nav():
    """上端に青い帯のグローバルナビゲーションを表示する（Streamlitネイティブ方式）"""
    st.markdown("""
        <style>
        /* Streamlit自体のレイアウト全体の上部余白を徹底的に削除 */
        .stApp, div[data-testid="stAppViewContainer"], section.main {
            padding-top: 0 !important;
            margin-top: 0 !important;
        }
        /* バージョン依存のすべてのメインコンテナをターゲットにする */
        .stMainBlockContainer, .stAppViewBlockContainer, div.block-container, section.main > div.block-container {
            padding-top: 0rem !important;
            margin-top: 0 !important;
        }
        /* デフォルトのヘッダーおよび装飾バーを完全に存在ごと消す */
        header[data-testid="stHeader"], div[data-testid="stDecoration"] {
            display: none !important;
            height: 0 !important;
        }
        #nav-band-marker { height: 0; padding: 0; margin: 0; overflow: hidden; }

        /* 擬似要素(::before)を使ってナビ行の背景に全幅の青帯を描画する（最も安全な方式） */
        div[data-testid="stHorizontalBlock"]:has(.nav-btn-target) {
            position: relative;
            padding-top: 8px !important;
            padding-bottom: 8px !important;
        }
        div[data-testid="stHorizontalBlock"]:has(.nav-btn-target)::before {
            content: "";
            position: absolute;
            top: 0; bottom: 0;
            left: -50vw; right: -50vw; /* 画面幅の半分ぶん左右に拡張して全幅ブルーにする */
            background-color: #1565c0;
            z-index: 0;
        }
        div[data-testid="stHorizontalBlock"]:has(.nav-btn-target) > div {
            position: relative;
            z-index: 1; /* コンテンツ類（ボタンなど）を青背景の手前に持ってくる */
        }

        /* ナビボタン：枠なし・白文字 */
        div[data-testid="stHorizontalBlock"]:has(.nav-btn-target) button,
        div[data-testid="stHorizontalBlock"]:has(.nav-btn-target) button[kind="secondary"] {
            background: transparent !important;
            border: none !important;
            color: white !important;
            box-shadow: none !important;
            outline: none !important;
            height: auto !important;
            min-height: 0 !important;
        }
        div[data-testid="stHorizontalBlock"]:has(.nav-btn-target) button p {
            color: white !important;
            font-size: 1rem !important;
            white-space: nowrap !important;
        }
        div[data-testid="stHorizontalBlock"]:has(.nav-btn-target) button:hover {
            text-decoration: underline !important;
            background: transparent !important;
            color: white !important;
        }
        /* シフト表編集（無効化ボタン）用の完全に同じレイアウトのスタイル */
        div[data-testid="stHorizontalBlock"]:has(.nav-btn-target) button:disabled {
            opacity: 0.55 !important;
            color: white !important;
            background: transparent !important;
            border: none !important;
            box-shadow: none !important;
            cursor: default !important;
        }
        div[data-testid="stHorizontalBlock"]:has(.nav-btn-target) button:focus,
        div[data-testid="stHorizontalBlock"]:has(.nav-btn-target) button:active {
            box-shadow: none !important;
            border: none !important;
            outline: none !important;
            background: transparent !important;
            color: white !important;
        }
        .nav-btn-target { display: none; }
        </style>
    """, unsafe_allow_html=True)

    # st.containerでナビバーをラップ
    with st.container():
        # マーカーのみ出力（背景はCSSの ::before で描画）
        st.markdown('<div id="nav-band-marker" style="height:0;padding:0;margin:0;overflow:hidden;"></div>', unsafe_allow_html=True)
        col1, col2, col3, col4, col5 = st.columns([1, 3, 2, 3, 1])
        with col1:
            st.markdown('<div class="nav-btn-target"></div>', unsafe_allow_html=True)
        with col2:
            if st.button("請求書送信", key="nav_claim", use_container_width=True):
                st.session_state.page = "claim_send"
                st.rerun()
        with col3:
            st.button("シフト表編集", disabled=True, key="nav_shift", use_container_width=True)
        with col4:
            if st.button("マイページ", key="nav_mypage", use_container_width=True):
                st.session_state.page = "mypage"
                st.rerun()


# データファイルのパス (ローカルテスト用。Streamlit Cloudではリポジトリ内パス)
MYPAGE_PATH = "My-page.xlsx"
SHIFT_PATH = "シフト表時計台警備通年.xlsx"
TEMPLATE_PATH = "テンプレート.xlsx"

# --- ユーティリティ関数 ---

def load_mypage():
    try:
        conn = st.connection("gsheets", type=GSheetsConnection)
        df = conn.read(worksheet="My-page", ttl=0) # ttl=0 で常に最新を取得
        
        # 必要な列の補填
        if "会社メアド" not in df.columns: df["会社メアド"] = ""
        if "pass" not in df.columns: df["pass"] = SHARED_PASSWORD
            
        return df
    except Exception:
        # スプレッドシートが取得できない場合は、フォールバックとしてローカルのファイルを読む
        if os.path.exists(MYPAGE_PATH):
            df = pd.read_excel(MYPAGE_PATH)
            if "会社メアド" not in df.columns: df["会社メアド"] = ""
            if "pass" not in df.columns: df["pass"] = SHARED_PASSWORD
            return df
        return pd.DataFrame(columns=["ID", "苗字", "氏名", "往復移動経路", "運賃", "送信", "メアド", "会社メアド", "pass"])

def save_mypage(df):
    try:
        conn = st.connection("gsheets", type=GSheetsConnection)
        conn.update(worksheet="My-page", data=df)
        return True
    except Exception as e:
        # フォールバック保存
        df.to_excel(MYPAGE_PATH, index=False)
        return False

def get_email_body(surname, year, month):
    template_path = "mail-message.txt"
    if os.path.exists(template_path):
        # 日本語環境特有のエンコーディングを考慮
        for enc in ['utf-8', 'cp932', 'shift_jis']:
            try:
                with open(template_path, 'r', encoding=enc) as f:
                    content = f.read()
                    # プレースホルダーの置換（全角・半角両方対応）
                    content = content.replace("{苗字}", surname).replace("｛苗字｝", surname)
                    year_month = f"{year}年{month}月"
                    content = content.replace("{年月}", year_month).replace("｛年月｝", year_month)
                    return content
            except:
                continue
    # テンプレートがない場合のフォールバック
    return f"高橋 様\n\n時計台警備の{surname}です。\nお疲れ様です。\n\n交通費請求書（{year}年{month}月分）をお送りします。\n\n以上、どうぞよろしくお願い致します。"

def load_shift_data(year, month):
    month_label = f"{month}月"
    if not os.path.exists(SHIFT_PATH):
        return pd.DataFrame(columns=["日付", "人名"])
    
    wb = openpyxl.load_workbook(SHIFT_PATH, data_only=True)
    target_ws = None
    start_row = None
    
    # 月の場所を探す
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in range(1, 100):
            for col in range(1, 10):
                if ws.cell(row=row, column=col).value == month_label:
                    target_ws = ws
                    start_row = row
                    break
            if target_ws: break
        if target_ws: break
    
    if not target_ws:
        return pd.DataFrame(columns=["日付", "人名"])

    # 日付列の特定
    day_cols = {}
    for c in range(1, target_ws.max_column + 1):
        val = target_ws.cell(row=start_row, column=c).value
        if isinstance(val, (int, float)) or (isinstance(val, str) and val.isdigit()):
            day_cols[int(val)] = c

    # 3人のデータを抽出
    data = []
    names = ["山口", "坂下", "堀川"]
    # 実際は全スタッフ分必要だが、今回はこの3人に注目
    for name in names:
        row_idx = None
        for r in range(start_row, start_row + 20):
            for c in range(1, 5):
                if target_ws.cell(row=r, column=c).value == name:
                    row_idx = r
                    break
            if row_idx: break
        
        if row_idx:
            for day, col in day_cols.items():
                val = target_ws.cell(row=row_idx, column=col).value
                # 空白を除去して判定（「朝 」「 夜」などに対応）
                if isinstance(val, str) and val.strip() in ["出", "朝", "夜"]:
                    data.append({"日付": day, "人名": name})
    
    # DataFrame化
    df = pd.DataFrame(data)
    
    # 正月三が日（1/1~1/3）以外は、1日1名に集約する（旧来の動作を維持）
    if not df.empty:
        if month == 1:
            # 1月の場合、4日以降は1日1名にする
            df_jan_top3 = df[df["日付"] <= 3]
            df_others = df[df["日付"] > 3].drop_duplicates(subset=["日付"], keep="last")
            df = pd.concat([df_jan_top3, df_others]).sort_values("日付")
        else:
            # 1月以外は全ての日で1日1名にする
            df = df.drop_duplicates(subset=["日付"], keep="last").sort_values("日付")
            
    return df

def save_shift_to_excel(year, month, edited_df):
    month_label = f"{month}月"
    wb = openpyxl.load_workbook(SHIFT_PATH)
    
    target_ws = None
    start_row = None
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in range(1, 100):
            for col in range(1, 10):
                if ws.cell(row=row, column=col).value == month_label:
                    target_ws = ws
                    start_row = row
                    break
            if target_ws: break
        if target_ws: break

    if not target_ws:
        return False

    day_cols = {}
    for c in range(1, target_ws.max_column + 1):
        val = target_ws.cell(row=start_row, column=c).value
        if isinstance(val, (int, float)) or (isinstance(val, str) and val.isdigit()):
            day_cols[int(val)] = c

    names = ["山口", "坂下", "堀川"]
    name_rows = {}
    for name in names:
        for r in range(start_row, start_row + 20):
            for c in range(1, 5):
                if target_ws.cell(row=r, column=c).value == name:
                    name_rows[name] = r
                    break
            if name in name_rows: break

    # 全員分のデータを一旦消去（指定月・指定名の範囲）
    for r in name_rows.values():
        for col in day_cols.values():
            target_ws.cell(row=r, column=col).value = None

    # 編集データを反映
    for _, row_data in edited_df.iterrows():
        day = row_data["日付"]
        name = row_data["人名"]
        if name in name_rows and day in day_cols:
            r = name_rows[name]
            col = day_cols[day]
            
            # 「出」を記入
            cell = target_ws.cell(row=r, column=col)
            cell.value = "出"
            cell.alignment = Alignment(horizontal="right")
            
            # 翌日を「明」にセット (ルール適用: 正月三が日は明けない)
            is_new_year = (month == 1 and day in [1, 2, 3])
            if not is_new_year and day + 1 in day_cols:
                ake_col = day_cols[day + 1]
                ake_cell = target_ws.cell(row=r, column=ake_col)
                ake_cell.value = "明"
                ake_cell.alignment = Alignment(horizontal="left")

    wb.save(SHIFT_PATH)
    return True

def generate_claim_pdf(user_id, month, year, route, fare, shift_data):
    # テンプレートPDFの読み込み
    if not os.path.exists("テンプレート.pdf"):
        # フォールバック処理（テンプレートがない場合）
        buffer = BytesIO()
        pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))
        c = canvas.Canvas(buffer, pagesize=A4)
        c.setFont('HeiseiKakuGo-W5', 20)
        c.drawString(100, 700, "テンプレート.pdf が見つかりません")
        c.save()
        buffer.seek(0)
        return buffer

    reader = PdfReader("テンプレート.pdf")
    template_page = reader.pages[0]
    
    # オーバーレイ作成（ReportLab）
    packet = BytesIO()
    pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))
    c = canvas.Canvas(packet, pagesize=A4)
    width, height = A4

    # 年月の記入 (動的に取得して上書き)
    c.setFont('HeiseiKakuGo-W5', 18)
    # 元の「2026年 2月分」という静的な文字を隠すための白い長方形
    c.setFillColorRGB(1, 1, 1) # 白
    c.rect(38, 789, 150, 25, fill=1, stroke=0)
    c.setFillColorRGB(0, 0, 0) # 黒に戻す
    c.drawString(38.88, 790.68, f"{year}年 {month}月分")

    # 氏名
    c.setFont('HeiseiKakuGo-W5', 12)
    c.drawString(205, 768, user_id) # 右に100pt、上に5pt移動 (105, 735 -> 205, 740)

    # テーブルの座標定義 (ユーザー指示に基づく最終調整)
    x_得意先_center = 96
    x_現場_center = 158
    x_経路_left = 194
    x_運賃_right = 450 # さらに5pt 右に移動 (445 -> 450)
    
    y_start = 719.64 # テンプレートの「1」のy座標 (+3pt)
    row_gap = 19.68 # 1日あたりの行間隔 (実測値)
    
    c.setFont('HeiseiKakuGo-W5', 10)
    
    # 月の日数を取得
    _, last_day = calendar.monthrange(year, month)
    
    total_fare = 0
    for day in range(1, last_day + 1):
        status = shift_data.get(day, "")
        if status == "出":
            curr_y = y_start - (day - 1) * row_gap
            
            # 得意先名
            c.drawCentredString(x_得意先_center, curr_y, "MMS")
            # 現場名
            c.drawCentredString(x_現場_center, curr_y, "時計台")
            
            # 往復移動経路 (長い場合はフォントを小さく)
            display_route = route
            font_size = 10
            if len(display_route) > 25: font_size = 7
            elif len(display_route) > 20: font_size = 8
            
            c.setFont('HeiseiKakuGo-W5', font_size)
            c.drawString(x_経路_left, curr_y, display_route)
            c.setFont('HeiseiKakuGo-W5', 10)
            
            # 運賃 (E列: 363〜430pt の範囲に修正)
            c.drawRightString(x_運賃_right, curr_y, f"{fare}")
            total_fare += fare
            
            # 走行km数と通信費は空欄のまま (要件)

    # 運賃の合計金額を記入 (計の行)
    y_total = 111.00 # テンプレートの「計」のy座標 (+3pt)
    c.drawRightString(x_運賃_right, y_total, f"{total_fare}")

    c.save()
    packet.seek(0)
    
    # 合成
    new_pdf_reader = PdfReader(packet)
    output = PdfWriter()
    # 注意: template_pageをコピーして使用
    import copy
    merged_page = copy.copy(template_page)
    merged_page.merge_page(new_pdf_reader.pages[0])
    output.add_page(merged_page)
    
    result_buffer = BytesIO()
    output.write(result_buffer)
    result_buffer.seek(0)
    return result_buffer

def send_email(to_email, subject, body, from_name="時計台警備", attachment_content=None, attachment_filename="file.pdf"):
    try:
        msg = MIMEMultipart()
        msg['From'] = formataddr((from_name, SENDER_EMAIL))
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))

        if attachment_content:
            part = MIMEApplication(attachment_content, Name=attachment_filename)
            part['Content-Disposition'] = f'attachment; filename="{attachment_filename}"'
            msg.attach(part)

        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(SENDER_EMAIL, EMAIL_PASSWORD)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        st.error(f"メール送信エラー: {e}")
        st.expander("詳細なエラー情報").code(error_msg)
        return False

# --- ページ構成 ---

def login_page():
    st.title(APP_TITLE)
    
    # URLパラメータからIDを取得 (?id=saka 等)
    query_params = st.query_params
    default_id = query_params.get("id", "")
    
    # セレクトボックス用の選択肢（表示名とIDの対応）
    id_options = {
        "山口": "yama",
        "坂下": "saka",
        "堀川": "hori"
    }
    # 反転させた辞書（IDから名前を引くため）
    id_to_name = {v: k for k, v in id_options.items()}
    
    # デフォルトのインデックスを決定
    default_index = 0
    if default_id in id_options.values():
        default_index = list(id_options.values()).index(default_id)

    with st.form("login"):
        # ID入力をセレクトボックスに変更
        selected_name = st.selectbox("お名前を選んでください", options=list(id_options.keys()), index=default_index)
        user_id = id_options[selected_name]
        
        password = st.text_input("合言葉を入力してください", type="password")
        if st.form_submit_button("ログイン"):
            # My-pageから個別パスワードを引いてくる
            df_mypage = load_mypage()
            user_row = df_mypage[df_mypage["ID"] == user_id]
            
            expected_pass = SHARED_PASSWORD
            if not user_row.empty:
                val = user_row.iloc[0].get("pass", "")
                if val and not (isinstance(val, float) and pd.isna(val)):
                    expected_pass = str(val)
            
            if user_id in VALID_IDS and password == expected_pass:
                st.session_state.logged_in = True
                st.session_state.user_id = user_id
                st.rerun()
            else:
                st.error("合言葉が正しくありません")

def get_jst_today():
    """UTCの実行環境でも現在の日本時間を返す"""
    # サーバーがUTCの場合、+9時間してJSTにする
    return (datetime.datetime.now(timezone.utc) + timedelta(hours=9)).date()

def shift_edit_page():
    render_global_nav()
    st.title("シフト表を編集")
    
    today = get_jst_today()
    month_options = [
        (today.replace(day=1) - datetime.timedelta(days=1)).strftime("%Y-%m"),
        today.strftime("%Y-%m"),
        (today.replace(day=1) + datetime.timedelta(days=32)).strftime("%Y-%m")
    ]
    selected_month_str = st.selectbox("年月を選択", month_options, index=1)
    year, month = map(int, selected_month_str.split("-"))
    
    # 実際のデータを読み込み
    df_shift = load_shift_data(year, month)
    if df_shift.empty:
        df_shift = pd.DataFrame({"日付": range(1, 32), "人名": "未設定"})
    
    st.write("人名を選択してください（出：右寄せ、明：左寄せに自動整形されます）")
    edited_df = st.data_editor(
        df_shift,
        column_config={
            "人名": st.column_config.SelectboxColumn(
                "担当者",
                options=["山口", "坂下", "堀川", "未設定"],
                required=True,
            )
        },
        disabled=[], # 複数名対応のため日付の編集も許可（または行追加可能にする）
        hide_index=True,
        num_rows="dynamic" if month == 1 else "fixed",
    )

    if st.button("変更を保存"):
        if save_shift_to_excel(year, month, edited_df):
            st.success("シフト表（Excel）を更新しました。")
            st.session_state.page = "claim_send"
            st.rerun()
        else:
            st.error("Excelの更新に失敗しました。対象の月が見つからない可能性があります。")




@st.dialog("送信確認")
def send_confirmation_dialog(user_info, year, month, pdf_buffer, filename_pdf, surname_label, email_body):
    st.markdown(f"""
        <style>
        /* ダイアログ内のボタンを確実に中央揃えにする */
        div[data-testid="stDialog"] button {{
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            height: 60px !important; /* 高さを固定して安定させる */
            font-size: 1.1rem !important;
        }}
        div[data-testid="stDialog"] button p {{
            margin: 0 !important;
            line-height: 1 !important;
        }}
        </style>
        <div style='text-align: center; padding: 10px 0 20px 0;'>
            <h3 style='margin-bottom: 25px; line-height: 1.6;'>
                会社と自分に、交通費請求書を<br>
                メール送信します。よろしいですか？
            </h3>
        </div>
    """, unsafe_allow_html=True)

    # 送信完了フラグを管理
    if "send_success" not in st.session_state:
        st.session_state.send_success = False

    if st.session_state.send_success:
        # 送信完了後の正方形フォーム表示
        st.markdown("""
            <style>
            .success-container {
                display: flex;
                flex-direction: column;
                align-items: center;
                justify-content: center;
                width: 66vw;
                aspect-ratio: 1 / 1;
                margin: 0 auto;
                background-color: white;
                border-radius: 10px;
                box-shadow: 0 4px 15px rgba(0,0,0,0.1);
                border: 1px solid #e0e0e0;
                padding: 20px;
                text-align: center;
            }
            .spinner-large {
                width: 60px;
                height: 60px;
                border: 6px solid #f3f3f3;
                border-top: 6px solid #4CAF50;
                border-radius: 50%;
                animation: spin 2s linear infinite;
                margin-bottom: 20px;
            }
            @keyframes spin {
                0% { transform: rotate(0deg); }
                100% { transform: rotate(360deg); }
            }
            .success-text {
                font-size: 1.5rem;
                font-weight: bold;
                color: #2e7d32;
                margin-bottom: 25px;
            }
            </style>
            <div class="success-container">
                <div class="spinner-large"></div>
                <div class="success-text">送信完了！</div>
            </div>
        """, unsafe_allow_html=True)
        if st.button("閉じる", use_container_width=True):
            st.session_state.send_success = False
            st.rerun()
        return

    col1, col2 = st.columns(2)
    with col1:
        if st.button("キャンセル", use_container_width=True):
            st.rerun()
    with col2:
        if st.button("送信", use_container_width=True, type="primary"):
            target_company_email = get_target_company_email()
                
            from_display_name = f"時計台警備（{user_info['氏名']}）"
            with st.spinner("送信中..."):
                success = send_email(
                    target_company_email, 
                    f"交通費請求書_{year}年{month}月_{surname_label}",
                    email_body,
                    from_name=from_display_name,
                    attachment_content=pdf_buffer.getvalue(),
                    attachment_filename=filename_pdf
                )
                if success:
                    try:
                        self_body = (
                            "以下の内容で、会社に交通費請求書を\n"
                            "送りました。\n\n"
                            "------------------------------\n"
                            f"{email_body}"
                        )
                        send_email(user_info["メアド"], "【送信済】交通費請求書", self_body, from_name=from_display_name, attachment_content=pdf_buffer.getvalue(), attachment_filename=filename_pdf)
                    except:
                        pass
                    st.session_state.send_success = True
                    st.rerun()
                else:
                    st.error("会社へのメール送信に失敗しました。")

def claim_send_page():
    render_global_nav()
    st.markdown("""
        <h1 style="text-align: center; line-height: 1.3;">
            交通費請求書を<br>表示・送信
        </h1>
    """, unsafe_allow_html=True)
    
    df_mypage = load_mypage()
    user_info = df_mypage[df_mypage["ID"] == st.session_state.user_id].iloc[0]
    user_name = user_info["氏名"] # 「氏名」カラムを使用するように変更
    
    # ユーザー名を表示する薄い青色の帯を追加（センター合わせ）
    st.markdown(f"""
        <style>
        .user-band {{
            background-color: #e3f2fd;
            color: #1565c0;
            text-align: center;
            padding: 4px 0 !important;
            border-radius: 4px;
            font-weight: bold;
            margin: 0 0 20px 0 !important; /* 上のタイトルとの隙間を詰め、下の要素と幅を合わせる */
            width: 100% !important; /* 下のカラム（col1+col2）の合計幅に合わせる */
            font-size: 0.9rem;
            border: 1px solid #bbdefb;
            display: block !important;
        }}
        .user-band p, .user-band div, .user-band span {{
            margin: 0 !important;
            padding: 0 !important;
            line-height: 1.2 !important; /* line-height を少し持たせて垂直中央を安定させる */
            display: inline-block !important;
        }}
        </style>
        <div class="user-band"><span>{user_name}さん</span></div>
    """, unsafe_allow_html=True)
    
    today = get_jst_today()
    first_of_this_month = today.replace(day=1)
    
    # 選択肢の生成（先月、今月、来月）
    options = []
    # 先月
    last_month_date = first_of_this_month - datetime.timedelta(days=1)
    options.append(last_month_date)
    # 今月
    options.append(first_of_this_month)
    # 来月
    next_month_date = (first_of_this_month + datetime.timedelta(days=32)).replace(day=1)
    options.append(next_month_date)
    
    # 表示用のラベル作成
    option_labels = [d.strftime("%Y年%m月").replace("年0", "年") for d in options]

    # --- レイアウト配置（モバイルでも横並びを維持） ---
    # パラメータ取得前に一旦計算する必要があるため、selectboxの直前に配置
    # (Streamlitの再実行特性を考慮し、セレクトボックスの値を先に確定させる)
    
    # セレクトボックスの状態を先に取得（またはデフォルト値を使用）
    temp_label = st.session_state.get("sel_resp_v11", option_labels[0])
    temp_date = options[option_labels.index(temp_label)]
    is_last_month = (temp_date == last_month_date)

    msg = """
    <div style="font-size: 0.8rem; line-height: 1.4; margin-bottom: 10px;">
    【使い方】<br>
    ●シフト表と照合し間違いないことを確認し、送信ボタン。<br>
    ●確認の赤い送信ボタンをクリック。送信完了！<br>
    ●送信できるのは先月分のみ。毎月1～5日の間に送信して下さい。<br>
    ●先月、今月、来月の請求内容が閲覧できます。<br>
    ●シフト表から勤務の日付を特定し、運賃等を書き込んでいます。
    </div>
    """
    st.markdown(msg, unsafe_allow_html=True)
    
    # CSS：縦積みを防ぎ、比率制御（data-testidに依存しない位置ターゲット）
    st.markdown("""
        <style>
        /* 横並びを常時維持 */
        [data-testid="stHorizontalBlock"]:has(.send-row-marker) {
            flex-direction: row !important;
            flex-wrap: nowrap !important;
            gap: 8px !important;
            align-items: flex-end !important; /* 絶対的な下揃え */
        }
        /* 55:45 比率でカラムを分割 */
        [data-testid="stHorizontalBlock"]:has(.send-row-marker) > div:nth-child(1) {
            flex: 55 1 0% !important;
            min-width: 0 !important;
            width: auto !important;
        }
        [data-testid="stHorizontalBlock"]:has(.send-row-marker) > div:nth-child(2) {
            flex: 45 1 0% !important;
            min-width: 0 !important;
            width: auto !important;
        }
        /* 高さを完全に一致させる（84px固定） */
        [data-testid="stHorizontalBlock"]:has(.send-row-marker) [data-testid="stSelectbox"] {
            margin: 0 !important;
            padding: 0 !important;
            display: flex !important;
            align-items: flex-end !important;
        }
        [data-testid="stHorizontalBlock"]:has(.send-row-marker) [data-testid="stSelectbox"] > div:first-child,
        [data-testid="stHorizontalBlock"]:has(.send-row-marker) [data-baseweb="select"] {
            height: 84px !important;
            min-height: 84px !important;
            margin: 0 !important;
        }
        [data-testid="stHorizontalBlock"]:has(.send-row-marker) [data-baseweb="select"] > div {
            height: 84px !important;
            min-height: 84px !important;
            align-items: center !important; /* 中央のテキスト（「2026年2月」）を上下中央に */
            border-radius: 8px !important;
        }
        [data-testid="stHorizontalBlock"]:has(.send-row-marker) button {
            height: 84px !important;
            min-height: 84px !important;
            border-radius: 8px !important;
            margin: 0 !important;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            overflow: hidden !important;
        }
        /* ボタン内部の要素（divやp）もスクロールバーが出ないように隠す */
        [data-testid="stHorizontalBlock"]:has(.send-row-marker) button > div,
        [data-testid="stHorizontalBlock"]:has(.send-row-marker) button p {
            overflow: hidden !important;
        }
        /* ボタンテキスト内の改行（\n）を有効化 */
        [data-testid="stHorizontalBlock"]:has(.send-row-marker) button p,
        [data-testid="stHorizontalBlock"]:has(.send-row-marker) button > div {
            white-space: pre-line !important;
            text-align: center !important;
            line-height: 1.4 !important;
        }
        /* 無効化時の「かすみ表示」設定 */
        [data-testid="stHorizontalBlock"]:has(.send-row-marker) button:disabled {
            opacity: 0.6 !important; /* 少し濃くして読みやすく */
            background-color: #f0f0f0 !important;
            color: #444 !important; /* 文字色を濃いグレーに変更 */
            border-color: #ccc !important;
            cursor: not-allowed !important;
        }
        /* セレクトボックスのラベルを非表示 */
        [data-testid="stWidgetLabel"] {
            display: none !important;
        }
        /* プレビュー画像：固定幅（実寸）表示。親コンテナの横スクロールも許可 */
        .pdf-preview-wrapper {
            overflow-x: auto !important;
            -webkit-overflow-scrolling: touch;
            width: 100% !important;
        }
        .pdf-preview-wrapper img {
            width: 794px !important;
            min-width: 794px !important;
            max-width: none !important;
            display: block;
        }
        /* markdownコンテナ自体がoverflowを隠している場合の対処 */
        [data-testid="stMarkdownContainer"] {
            overflow-x: auto !important;
        }
        /* 先月判定マーカーがある場合のボタン装飾 */
        div[data-testid="stHorizontalBlock"]:has(.blue-btn-marker) button {
            background-color: #e3f2fd !important;
            color: #1565c0 !important;
            border: 1px solid #90caf9 !important;
            font-weight: bold;
        }
        .blue-btn-marker {
            display: none !important;
        }
        /* ボタンが1行になり、かつ背景色がある場合にテキストの中央揃えを強調 */
        div[data-testid="stHorizontalBlock"]:has(.blue-btn-marker) button p {
            line-height: 1.2 !important;
        }
        /* 高さを狂わせる原因となるマーカー要素のコンテナ余白を完全に消滅させる */
        div[data-testid="element-container"]:has(.send-row-marker) {
            display: none !important;
            margin: 0 !important;
            padding: 0 !important;
            height: 0 !important;
        }
        </style>
    """, unsafe_allow_html=True)
    
    # カラム比率 2:1（幅に余裕をもたせてギャップ込みでも収まるように）
    col1, col2 = st.columns([11, 9])
    
    with col1:
        # マーカー要素を1つにまとめて出力し、CSSでコンテナごと隠す
        marker_html = '<div class="send-row-marker"></div>'
        if is_last_month:
            marker_html += '<div class="blue-btn-marker"></div>'
        st.markdown(marker_html, unsafe_allow_html=True)
        
        selected_label = st.selectbox("年月", options=option_labels, index=0, label_visibility="collapsed", key="sel_resp_v11")
    
    # 2. データの準備
    selected_date = options[option_labels.index(selected_label)]
    year, month = selected_date.year, selected_date.month
    
    # 先月かどうかの判定 (CSSマーカー用)
    is_last_month = (selected_date == last_month_date)
    
    # --- 送信制限ロジック：今月・未来の分は送信不能にする ---
    today_jst = get_jst_today()
    # 選択された年月の初日と、今月の初日を比較
    first_of_selected = datetime.date(year, month, 1)
    first_of_current_month = today_jst.replace(day=1)
    is_future_or_current = first_of_selected >= first_of_current_month
    
    # 名字ラベルの取得（苗字がなければ氏名の先頭を使用）
    surname_label = user_info.get("苗字", "")
    if not surname_label or (isinstance(surname_label, float) and pd.isna(surname_label)):
        surname_label = user_name.split()[0] if user_name else "請求書"
        
    filename_pdf = f"交通費請求書_{year}_{month}_{surname_label}.pdf"

    id_to_name = {"yama": "山口", "saka": "坂下", "hori": "堀川"}
    kanji_name = id_to_name.get(st.session_state.user_id, "堀川")
    raw_shift = load_shift_data(year, month)
    
    # ログインユーザーのシフトのみを抽出（1日複数名対応）
    personal_shift = {}
    if not raw_shift.empty:
        user_records = raw_shift[raw_shift["人名"] == kanji_name]
        personal_shift = {row["日付"]: "出" for _, row in user_records.iterrows()}
    
    pdf_buffer = generate_claim_pdf(user_name, month, year, user_info["往復移動経路"], user_info["運賃"], personal_shift)
    
    with col2:
        btn_label = "会社と自分に\n送　信"
        if is_future_or_current:
            btn_label = "今月以降は\n送信不可"
            
        if st.button(btn_label, use_container_width=True, key="btn_final_v2", disabled=is_future_or_current):
            email_body = get_email_body(surname_label, year, month)
            send_confirmation_dialog(user_info, year, month, pdf_buffer, filename_pdf, surname_label, email_body)

    # --- メッセージプレビューの表示 ---
    target_company_email = get_target_company_email()
    st.markdown(f"""<div style="font-size: 0.8rem; background-color: #f9f9f9; padding: 5px 10px; border: 1px solid #ddd; border-radius: 5px; margin-top: 5px; margin-bottom: 5px; white-space: pre-wrap;">【メッセージ】
高橋　様　{target_company_email}
時計台警備の{surname_label}です。お疲れ様です。
交通費請求用紙　お送りします。
{year}年{month}月分です。
以上、どうぞ、よろしくお願い致します。</div>""", unsafe_allow_html=True)

    # --- プレビューとダウンロード ---
    def pdf_to_image(pdf_bytes):
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        page = doc.load_page(0)
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
        return pix.tobytes("png")

    image_bytes = pdf_to_image(pdf_buffer.getvalue())
    # 画像を固定幅（実寸）で表示 — スマホで縮小させない
    import base64
    img_b64 = base64.b64encode(image_bytes).decode()
    st.markdown(f"""
        <div class="pdf-preview-wrapper">
            <img src="data:image/png;base64,{img_b64}"
                 alt="請求書プレビュー（確認用）">
        </div>
        <p style="font-size: 0.8em; color: #888; margin-top: 4px;">請求書プレビュー（確認用）</p>
    """, unsafe_allow_html=True)

    st.download_button(
        label="PDFをダウンロードして確認",
        data=pdf_buffer,
        file_name=filename_pdf,
        mime="application/pdf"
    )


def mypage_page():
    render_global_nav()
    st.title("マイページ")
    df_mypage = load_mypage()
    user_row = df_mypage[df_mypage["ID"] == st.session_state.user_id]
    
    if user_row.empty:
        # 新規登録（初期値）
        new_row = {"ID": st.session_state.user_id, "苗字": "", "氏名": "", "往復移動経路": "", "運賃": 460, "送信": "手動", "メアド": ""}
    if user_row.empty:
        # 新規登録（初期値）
        new_row = {"ID": st.session_state.user_id, "苗字": "", "氏名": "", "往復移動経路": "", "運賃": 460, "送信": "手動", "メアド": "", "会社メアド": "", "pass": SHARED_PASSWORD}
        df_mypage = pd.concat([df_mypage, pd.DataFrame([new_row])], ignore_index=True)
        user_row = df_mypage[df_mypage["ID"] == st.session_state.user_id]

    with st.form("mypage_form"):
        full_name = st.text_input("氏名", value=user_row.iloc[0]["氏名"])
        
        # 追加：パスワード設定
        current_pass = user_row.iloc[0].get("pass", SHARED_PASSWORD)
        if isinstance(current_pass, float) and pd.isna(current_pass):
            current_pass = SHARED_PASSWORD
        user_pass = st.text_input("個別の合言葉（パスワード）", value=str(current_pass), type="password")
        
        email = st.text_input("メールアドレス", value=user_row.iloc[0]["メアド"])
        route = st.text_input("往復移動経路", value=user_row.iloc[0]["往復移動経路"])
        fare = st.number_input("往復運賃", value=int(user_row.iloc[0]["運賃"]))
        send_setting = st.radio("送信設定", options=["手動", "毎月1日に自動"], index=0 if user_row.iloc[0]["送信"] == "手動" else 1)
        
        # 会社メアド（送信先）: 堀川氏のみ変更可能
        is_hori = st.session_state.user_id == "hori"
        company_email_val = st.selectbox(
            "会社メアド（送信先）", 
            options=email_options, 
            index=email_options.index(current_val) if current_val in email_options else 0,
            disabled=not is_hori,
            help="※送信先の設定は管理者のみ変更可能です" if not is_hori else None
        )
        
        # 保存ボタンのスタイルカスタマイズ
        st.markdown("""
            <style>
            /* マイページの保存ボタンを装飾（複数セレクタで確実に適用） */
            div[data-testid="stFormSubmitButton"] > button,
            div[data-testid="stForm"] button[type="submit"] {
                background-color: #ffb6c1 !important; /* 薄いピンク */
                color: #333333 !important;
                padding: 12px 40px !important; /* 天地1.5倍、左右2倍 */
                font-size: 1.2rem !important;
                font-weight: bold !important;
                border-radius: 8px !important;
                border: none !important;
                width: auto !important;
                margin: 20px auto !important;
                display: block !important;
                transition: all 0.3s ease !important;
                box-shadow: 0 4px 15px rgba(255, 182, 193, 0.4) !important;
            }
            div[data-testid="stFormSubmitButton"] > button:hover,
            div[data-testid="stForm"] button[type="submit"]:hover {
                background-color: #ff69b4 !important; /* ホバー時は少し濃く */
                color: white !important;
                transform: scale(1.05) !important;
                box-shadow: 0 6px 20px rgba(255, 105, 180, 0.4) !important;
            }
            </style>
        """, unsafe_allow_html=True)
        
        if st.form_submit_button("保存"):
            idx = df_mypage[df_mypage["ID"] == st.session_state.user_id].index[0]
            # 内部的なファイル名用ラベルとして苗字も更新しておく（氏名の最初の単語）
            if full_name:
                df_mypage.at[idx, "苗字"] = full_name.split()[0]
            df_mypage.at[idx, "氏名"] = full_name
            df_mypage.at[idx, "往復移動経路"] = route
            df_mypage.at[idx, "運賃"] = fare
            df_mypage.at[idx, "送信"] = "1日に自動" if send_setting == "毎月1日に自動" else "手動"
            df_mypage.at[idx, "メアド"] = email
            df_mypage.at[idx, "pass"] = user_pass
            df_mypage.at[idx, "会社メアド"] = company_email_val
            save_mypage(df_mypage)
            st.success("マイページを更新しました")

    # --- 管理者専用機能 ---
    if st.session_state.user_id == "hori":
        st.markdown("---")
        st.subheader("⚙️ 管理者専用機能")
        st.info("現在クラウド上で稼働している最新のユーザーデータ（Googleスプレッドシート）をExcel形式でダウンロードできます。")
        try:
            # 最新データを取得
            df_latest = load_mypage()
            
            # メモリ上でExcelファイルを作成
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_latest.to_excel(writer, index=False, sheet_name='My-page')
            excel_data = output.getvalue()

            st.download_button(
                label="最新のユーザーデータ (My-page) をダウンロード",
                data=excel_data,
                file_name=f"最新_My-page_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        except Exception as e:
            st.error(f"データの取得または変換に失敗しました: {e}")



# --- メインロジック ---

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "page" not in st.session_state:
    st.session_state.page = "claim_send"

if not st.session_state.logged_in:
    login_page()
else:
    if st.session_state.page == "shift_edit":
        shift_edit_page()
    elif st.session_state.page == "claim_send":
        claim_send_page()
    elif st.session_state.page == "mypage":
        mypage_page()
