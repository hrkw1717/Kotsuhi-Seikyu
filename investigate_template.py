import openpyxl

path = r"C:\Users\sbs\Documents\Antigravity\Tokeidai\テンプレート.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

print(f"Sheet names: {wb.sheetnames}")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- Sheet: {sheet_name} ---")
    # Print first few rows
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        print(row)
