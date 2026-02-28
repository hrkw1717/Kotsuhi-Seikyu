import openpyxl

path = r"C:\Users\sbs\Documents\Antigravity\Tokeidai\My-page.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- Sheet: {sheet_name} ---")
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        print(row)
