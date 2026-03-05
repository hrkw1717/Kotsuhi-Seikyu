import os
import sys
import calendar
import openpyxl
import pandas as pd
import json
from google.oauth2.service_account import Credentials
import gspread
from dotenv import load_dotenv

# プロジェクトルートディレクトリ
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# バックエンドディレクトリ（.envがある場所）
BACKEND_DIR = os.path.join(BASE_DIR, "backend")

# 環境変数の読み込み (.env はルートまたはbackendにある想定)
dotenv_path = os.path.join(BASE_DIR, ".env")
if not os.path.exists(dotenv_path):
    dotenv_path = os.path.join(BACKEND_DIR, ".env")
load_dotenv(dotenv_path)

SPREADSHEET_URL = os.getenv("SPREADSHEET_URL")
EXCEL_PATH = os.path.join(BASE_DIR, "シフト表時計台警備通年.xlsx")
SHEET_NAME_EXCEL = "2601-2612"

STAFF_NAMES = ["山口", "堀川", "坂下"]

def get_gsheet_client():
    creds_json = os.getenv("GCP_SERVICE_ACCOUNT_JSON")
    if not creds_json:
        raise ValueError("GCP_SERVICE_ACCOUNT_JSON is not set in .env")
    
    scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    creds_info = json.loads(creds_json)
    creds = Credentials.from_service_account_info(creds_info, scopes=scopes)
    return gspread.authorize(creds)

def fetch_spreadsheet_data():
    """スプレッドシートの「シフト表」から全データを取得"""
    print("⏳ スプレッドシートからデータを取得中...")
    client = get_gsheet_client()
    sh = client.open_by_url(SPREADSHEET_URL)
    worksheet = sh.worksheet("シフト表")
    return worksheet.get_all_records()

def sync_to_excel():
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ エラー: Excelファイルが見つかりません: {EXCEL_PATH}")
        return

    # 1. スプレッドシートの全データを取得
    try:
        ss_data = fetch_spreadsheet_data()
    except Exception as e:
        print(f"❌ スプレッドシートの取得に失敗しました: {e}")
        return
        
    df_ss = pd.DataFrame(ss_data)
    if "Year" not in df_ss.columns or "Month" not in df_ss.columns:
        print("❌ スプレッドシートの構造が想定と異なります（Year, Month列がありません）")
        return

    # 2. ローカルのExcelを開く
    print(f"⏳ Excelファイルを読み込み中: {os.path.basename(EXCEL_PATH)}")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    if SHEET_NAME_EXCEL not in wb.sheetnames:
        print(f"❌ Excelにシート '{SHEET_NAME_EXCEL}' が見つかりません")
        return
        
    ws = wb[SHEET_NAME_EXCEL]
    
    # 3. 各月(1〜12月)のブロック位置（行）を特定
    month_blocks = {}
    for m in ws.merged_cells.ranges:
        if m.min_col == 3 and m.max_col == 3:  # C列
            cell_val = ws.cell(row=m.min_row, column=3).value
            if cell_val and isinstance(cell_val, str) and '月' in cell_val:
                month_num = cell_val.replace('月', '').strip()
                if month_num.isdigit():
                    month_blocks[int(month_num)] = (m.min_row, m.max_row)

    if not month_blocks:
        print("❌ 月名ブロック(C列の結合セル)が見つかりませんでした")
        return

    print("✅ Excelの月ブロックを検出しました")

    # 4. 年を指定（現在は2026年固定）
    target_year = 2026
    update_count = 0

    # 5. 各月ごとに処理
    for month in range(1, 13):
        if month not in month_blocks:
            continue
            
        r_start, r_end = month_blocks[month]
        
        # スプレッドシートから該当する年月のデータを抽出
        df_month = df_ss[(df_ss["Year"] == target_year) & (df_ss["Month"] == month)]
        if df_month.empty:
            continue
            
        print(f"🔄 {month}月のデータを同期中...")
        last_day = calendar.monthrange(target_year, month)[1]
        
        # 各スタッフの行を見つける (ブロックの前後を探索)
        staff_rows = {}
        for row_idx in range(max(1, r_start - 2), min(ws.max_row + 1, r_end + 5)):
            c_val = ws.cell(row=row_idx, column=3).value
            if c_val in STAFF_NAMES:
                staff_rows[c_val] = row_idx
        
        # スタッフごとにデータを上書き
        for name in STAFF_NAMES:
            if name not in staff_rows:
                continue
                
            row_idx = staff_rows[name]
            ss_user_row = df_month[df_month["氏名"] == name]
            
            if ss_user_row.empty:
                continue
                
            sr = ss_user_row.iloc[0]
            
            # 1日〜月末日までの値をD列(4列目)以降にセット
            for day in range(1, last_day + 1):
                col_idx = 3 + day  # D列=4列目 (3 + 1)
                
                # スプレッドシートの値
                val = sr.get(str(day), "")
                
                # スプレッドシートとExcelの値を比較・更新
                current_cell = ws.cell(row=row_idx, column=col_idx)
                
                # None は空文字として扱う
                val_to_write = str(val).strip() if pd.notna(val) and str(val).strip() else None
                current_val_str = str(current_cell.value).strip() if current_cell.value is not None else ""
                val_to_write_str = val_to_write if val_to_write is not None else ""
                
                if current_val_str != val_to_write_str:
                    current_cell.value = val_to_write
                    update_count += 1
            
            # 勤務日数の合計列（AH列付近など）は既存の計算式か値を維持（今回はシフト記号のみ更新）

    # 6. 保存
    if update_count > 0:
        print(f"💾 {update_count}セルの変更を保存しています...")
        # ファイルがロックされている場合のエラーハンドリング
        try:
            wb.save(EXCEL_PATH)
            print("🌟 同期完了！Excelファイルが更新されました。")
        except PermissionError:
            print(f"❌ エラー: Excelファイルが開かれています。閉じてからもう一度実行してください ({EXCEL_PATH})")
    else:
        print("✨ 変更箇所はありませんでした (最新状態です)")

if __name__ == "__main__":
    print(f"--- 時計台警備 シフト表 同期ツール ---")
    sync_to_excel()
