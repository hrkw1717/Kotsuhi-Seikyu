import openpyxl

def check_horikawa_shifts(year, month):
    path = "シフト表時計台警備通年.xlsx"
    month_label = f"{month}月"
    wb = openpyxl.load_workbook(path, data_only=True)
    
    target_ws = None
    start_row = None
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in range(1, 100):
            for col in range(1, 10):
                if ws.cell(row=row, column=col).value == month_label:
                    target_ws = ws
                    start_row = row
                    break
            if target_ws: break
        if target_ws: break
    
    if not target_ws:
        print(f"{month_label} not found")
        return

    day_cols = {}
    for c in range(1, target_ws.max_column + 1):
        val = target_ws.cell(row=start_row, column=c).value
        # valが数値または数値文字列の場合
        if isinstance(val, (int, float)) or (isinstance(val, str) and val.strip().isdigit()):
            day_cols[int(val)] = c

    horikawa_row = None
    for r in range(start_row, start_row + 20):
        for c in range(1, 5):
            if target_ws.cell(row=r, column=c).value == "堀川":
                horikawa_row = r
                break
        if horikawa_row: break

    if not horikawa_row:
        print("堀川 not found")
        return

    work_days = []
    for day in range(1, 32):
        if day in day_cols:
            val = target_ws.cell(row=horikawa_row, column=day_cols[day]).value
            if val == "出":
                work_days.append(day)
    
    print(f"2026年{month}月 堀川さんの「出」の日: {work_days}")

if __name__ == "__main__":
    check_horikawa_shifts(2026, 2)
